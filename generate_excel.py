"""
Dastur 18-24: menejerlar haftalik sotuvi → Excel (oq tema)
"""
import urllib.request, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_URL = "https://docs.google.com/spreadsheets/d/e/2PACX-1vTLxosD6vFsi8LRgiPGUPlToQ-wGrkfcCJ9l8oI1y7RULkaA0_N9b9Uq6NDtNXzfgs_oBooKtFPGyzX/pub?gid=1003386002&single=true&output=csv"
OUTPUT  = "/Users/mac/Work Milliard/Dastur 24/Dastur_Manager_Sotuv.xlsx"

# ── RANGLAR (oq tema) ────────────────────────────────────────────────────
BG       = "FFFFFF"   # oq fon
H1       = "1D4ED8"   # to'q ko'k — asosiy sarlavha fon
H1_TXT   = "FFFFFF"
H2       = "DBEAFE"   # och ko'k — hafta header
H2_TXT   = "1E3A8A"
H3       = "EFF6FF"   # juda och ko'k — dastur sub-header
H3_TXT   = "1D4ED8"
ROW_ALT  = "F8FAFF"   # alternativ qator
TOTAL_BG = "DBEAFE"   # jami qator
TOTAL_TXT= "1E3A8A"
BORDER_C = "CBD5E1"   # border rangi
TEXT     = "1E293B"   # asosiy matn
MUTED    = "94A3B8"   # kulrang (nol qiymat)
GOOD     = "059669"   # yashil (yuqori son)

# Dastur ranglari (sub-header uchun)
D_COLORS = {
    "18": ("F3F4F6","374151"),  # kulrang
    "19": ("CCFBF1","0F766E"),  # teal
    "20": ("EDE9FE","6D28D9"),  # binafsha
    "21": ("FEF9C3","854D0E"),  # sariq
    "22": ("FFEDD5","9A3412"),  # to'q sariq
    "23": ("DCFCE7","166534"),  # yashil
    "24": ("DBEAFE","1E40AF"),  # ko'k
}

def fill(c): return PatternFill("solid", fgColor=c)
def font(bold=False, color=TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def border():
    s = Side(style="thin", color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border():
    t = Side(style="medium", color="9CA3AF")
    n = Side(style="thin",   color=BORDER_C)
    return Border(left=t, right=t, top=t, bottom=t)
def al(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cell(ws, row, col, value=None, **kw):
    c = ws.cell(row, col, value=value)
    if "fill"   in kw: c.fill      = kw["fill"]
    if "font"   in kw: c.font      = kw["font"]
    if "border" in kw: c.border    = kw["border"]
    if "align"  in kw: c.alignment = kw["align"]
    if "fmt"    in kw: c.number_format = kw["fmt"]
    return c

# ── CSV PARSER ───────────────────────────────────────────────────────────
def split_csv(line):
    res, cur, inq = [], "", False
    for ch in line:
        if ch == '"': inq = not inq; continue
        if ch == ',' and not inq: res.append(cur.strip()); cur = ""; continue
        cur += ch
    res.append(cur.strip()); return res

def parse_date(s):
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s.strip())
    return f"{m[3]}-{m[2].zfill(2)}-{m[1].zfill(2)}" if m else None

def to_int(s):
    try: return int(str(s).replace(" ","").replace("\xa0",""))
    except: return 0

def fetch():
    with urllib.request.urlopen(CSV_URL) as r:
        text = r.read().decode("utf-8")
    grid = [split_csv(l) for l in text.split("\n")]
    dasturlar, i = [], 0
    while i < len(grid):
        row = grid[i]
        cell_val = (row[2] if len(row)>2 else "") + (row[0] if row else "")
        dm = re.search(r"(\d{2})-?Дастур", cell_val)
        if not dm: i += 1; continue
        name = dm.group(0); i += 1
        starts, ends, wk_cnt, bcol, ocol = [], [], 0, -1, -1
        while i < len(grid):
            r = grid[i]; r0 = r[0].strip() if r else ""; r1 = r[1].strip() if len(r)>1 else ""
            if r0 == "Boshlanish sana": starts = [parse_date(c) for c in r[2:] if parse_date(c)]
            elif r0 == "Tugash sana":   ends   = [parse_date(c) for c in r[2:] if parse_date(c)]
            elif r1 in ("жами","жами") and len(r)>2 and "Hafta" in r[2]:
                wk_cnt = sum(1 for c in r[2:] if "Hafta" in c)
                for ci,c in enumerate(r):
                    if "бошқа" in c.lower(): bcol = ci
                    if "отказ" in c.lower(): ocol = ci
                i += 1; break
            elif len(r)>2 and re.search(r"\d{2}-?Дастур", r[2]) and not r0: break
            i += 1
        if wk_cnt == 0: continue
        weeks = [{"index":w+1,"label":f"{w+1}-hafta",
                  "start":starts[w] if w<len(starts) else None,
                  "end":  ends[w]   if w<len(ends)   else None}
                 for w in range(wk_cnt)]
        managers, total_row = [], None
        while i < len(grid):
            r = grid[i]; r0 = r[0].strip() if r else ""
            if len(r)>2 and re.search(r"\d{2}-?Дастур", r[2]) and not r0: break
            if re.match(r"^(жами:|Jami:|jami:)$", r0, re.I):
                wv = [to_int(r[2+w]) if 2+w<len(r) else 0 for w in range(wk_cnt)]
                total_row = {"total":to_int(r[1]) if len(r)>1 else 0, "weeks":wv,
                             "boshqa": to_int(r[bcol]) if bcol>=0 and bcol<len(r) else 0,
                             "otkazlar": to_int(r[ocol]) if ocol>=0 and ocol<len(r) else 0}
                i += 1; break
            if r0 and not re.match(r"^(Boshlanish|Tugash|жами|Jami)", r0, re.I):
                wv = [to_int(r[2+w]) if 2+w<len(r) else 0 for w in range(wk_cnt)]
                managers.append({"name":r0,"total":to_int(r[1]) if len(r)>1 else 0,
                                 "weeks":wv,
                                 "boshqa": to_int(r[bcol]) if bcol>=0 and bcol<len(r) else 0,
                                 "otkazlar": to_int(r[ocol]) if ocol>=0 and ocol<len(r) else 0})
            i += 1
        dasturlar.append({"name":name,"num":re.search(r"\d+",name).group(),
                          "weeks":weeks,"managers":managers,"total_row":total_row})
    dasturlar.sort(key=lambda d: int(d["num"]))
    return dasturlar

# ── PIVOT SHEET: 1-hafta → dasturlar → menejer ──────────────────────────
def write_pivot(wb, dasturlar, all_managers):
    ws = wb.create_sheet("HAFTA × DASTUR × MENEJER", 0)
    ws.sheet_view.showGridLines = False
    ws.tab_color = "1D4ED8"

    max_wk = max(len(d["weeks"]) for d in dasturlar)
    n_d    = len(dasturlar)
    # Ustunlar: A=Menejer, B=Grand Total, C onward: hafta*dastur
    # Har bir hafta uchun n_d ta ustun + 1 ta jami ustun
    # Layout: A | B(total) | [hafta1: D18,D19..D24, sum] | [hafta2: ...] | ...
    # Aslida foydalanuvchi "hafta asosiy column, dasturlar tagida" dedi
    # Shunday qilamiz: har hafta uchun n_d ta ustun (dasturlar), keyin "Hafta jami"

    # ── Kengliklari ──
    ws.column_dimensions["A"].width = 24   # Menejer
    ws.column_dimensions["B"].width = 9    # Grand Total
    for hf in range(max_wk):
        for di in range(n_d):
            col = 3 + hf*(n_d+1) + di
            ws.column_dimensions[get_column_letter(col)].width = 5
        # Hafta jami ustun
        sum_col = 3 + hf*(n_d+1) + n_d
        ws.column_dimensions[get_column_letter(sum_col)].width = 6

    # ── ROW 1: Sarlavha ──
    ws.row_dimensions[1].height = 30
    total_cols = 2 + max_wk*(n_d+1)
    c = ws.cell(1,1,"HAFTA × DASTUR × MENEJER SOTUV JADVALI")
    c.fill=fill(H1); c.font=font(bold=True,size=14,color=H1_TXT); c.alignment=al("left"); c.border=border()
    for col in range(2, total_cols+1):
        ws.cell(1,col).fill=fill(H1); ws.cell(1,col).border=border()
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_cols)

    # ── ROW 2: Hafta asosiy headerlar (merged n_d+1 ta ustun) ──
    ws.row_dimensions[2].height = 22
    # A2: "Menejer"
    c=ws.cell(2,1,"Menejer"); c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=10); c.alignment=al("left"); c.border=border()
    ws.merge_cells(start_row=2,start_column=1,end_row=3,end_column=1)
    # B2: "Grand Total"
    c=ws.cell(2,2,"Grand\nTotal"); c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=9); c.alignment=al("center",wrap=True); c.border=border()
    ws.merge_cells(start_row=2,start_column=2,end_row=3,end_column=2)

    for hf in range(max_wk):
        first_col = 3 + hf*(n_d+1)
        last_col  = first_col + n_d  # +1 uchun hafta jami
        # Hafta header (merged)
        c=ws.cell(2, first_col, f"{hf+1} - Hafta")
        c.fill=fill(H2); c.font=font(bold=True,color=H2_TXT,size=10); c.alignment=al("center"); c.border=border()
        for col in range(first_col+1, last_col+1):
            ws.cell(2,col).fill=fill(H2); ws.cell(2,col).border=border()
        ws.merge_cells(start_row=2,start_column=first_col,end_row=2,end_column=last_col)

    # ── ROW 3: Dastur sub-headerlar ──
    ws.row_dimensions[3].height = 18
    for hf in range(max_wk):
        for di, d in enumerate(dasturlar):
            col = 3 + hf*(n_d+1) + di
            dc, dt = D_COLORS.get(d["num"], ("EFF6FF","1D4ED8"))
            c=ws.cell(3,col,d["num"])
            c.fill=fill(dc); c.font=font(bold=True,color=dt,size=9); c.alignment=al("center"); c.border=border()
        # Hafta jami sub-header
        sum_col = 3 + hf*(n_d+1) + n_d
        c=ws.cell(3,sum_col,"Σ")
        c.fill=fill(H2); c.font=font(bold=True,color=H2_TXT,size=9); c.alignment=al("center"); c.border=border()

    # ── Sort managers by grand total ──
    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"]==name), None)
            if m: t += sum(m["weeks"])
        return t
    sorted_mgr = sorted(all_managers, key=grand_total, reverse=True)

    # ── ROW 4+: Menejer datalari ──
    first_data = 4
    for ri, name in enumerate(sorted_mgr):
        row = first_data + ri
        ws.row_dimensions[row].height = 17
        bg = BG if ri%2==0 else ROW_ALT

        # Menejer nomi
        gt = grand_total(name)
        c=ws.cell(row,1,name)
        c.fill=fill(bg); c.font=font(bold=gt>0,color=TEXT if gt>0 else MUTED,size=10)
        c.alignment=al("left"); c.border=border()

        # Grand Total formula
        # Grand total = sum of all "hafta jami" ustunlari
        sum_cols = [get_column_letter(3 + hf*(n_d+1) + n_d) for hf in range(max_wk)]
        formula = "=" + "+".join(f"{cl}{row}" for cl in sum_cols)
        gc=ws.cell(row,2,formula)
        gc.fill=fill(TOTAL_BG if gt>0 else bg)
        gc.font=font(bold=True,color=TOTAL_TXT if gt>0 else MUTED,size=11)
        gc.alignment=al("center"); gc.border=thick_border(); gc.number_format="0"

        # Hafta × Dastur values
        for hf in range(max_wk):
            hafta_vals_cols = []
            for di, d in enumerate(dasturlar):
                col = 3 + hf*(n_d+1) + di
                m   = next((x for x in d["managers"] if x["name"]==name), None)
                v   = m["weeks"][hf] if m and hf < len(m["weeks"]) else None

                c = ws.cell(row, col)
                if v is None:
                    # Bu dasturda bu hafta yo'q
                    c.value = None
                    c.fill  = fill("F1F5F9")
                    c.font  = font(color="E2E8F0",size=9)
                elif v == 0:
                    c.value = None
                    c.fill  = fill(bg)
                    c.font  = font(color="CBD5E1",size=9)
                else:
                    c.value = v
                    dc_bg, dc_txt = D_COLORS.get(d["num"],("DBEAFE","1E40AF"))
                    # Intensity bo'yicha to'qroq rang
                    mx = max((x["weeks"][hf] if hf<len(x["weeks"]) else 0) for x in d["managers"]) or 1
                    if v == mx and v > 1:
                        c.fill = fill("1D4ED8"); c.font = font(bold=True,color="FFFFFF",size=10)
                    elif v >= mx*0.6:
                        c.fill = fill("3B82F6"); c.font = font(bold=True,color="FFFFFF",size=10)
                    elif v >= mx*0.3:
                        c.fill = fill(dc_bg); c.font = font(bold=True,color=dc_txt,size=10)
                    else:
                        c.fill = fill(H3); c.font = font(color=H3_TXT,size=10)
                c.alignment = al("center"); c.border = border(); c.number_format="0"
                hafta_vals_cols.append(get_column_letter(col))

            # Hafta jami (formula)
            sum_col = 3 + hf*(n_d+1) + n_d
            rng = "+".join(f"{cl}{row}" for cl in hafta_vals_cols)
            sc=ws.cell(row, sum_col, f"={rng}")
            hf_total = sum(
                (next((x for x in d["managers"] if x["name"]==name),None) or {"weeks":[]})["weeks"][hf]
                if hf < len((next((x for x in d["managers"] if x["name"]==name),None) or {"weeks":[]})["weeks"]) else 0
                for d in dasturlar
            )
            sc.fill=fill(H2 if hf_total>0 else ROW_ALT)
            sc.font=font(bold=hf_total>0, color=H2_TXT if hf_total>0 else MUTED, size=10)
            sc.alignment=al("center"); sc.border=thick_border(); sc.number_format="0"

    # ── JAMI QATORI ──
    total_row = first_data + len(sorted_mgr)
    ws.row_dimensions[total_row].height = 22
    fr = first_data; lr = total_row - 1

    c=ws.cell(total_row,1,"JAMI")
    c.fill=fill(TOTAL_BG); c.font=font(bold=True,color=TOTAL_TXT,size=11); c.alignment=al("left"); c.border=border()

    c=ws.cell(total_row,2,f"=SUM(B{fr}:B{lr})")
    c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=12); c.alignment=al("center"); c.border=thick_border(); c.number_format="0"

    for hf in range(max_wk):
        for di in range(n_d):
            col=3+hf*(n_d+1)+di
            cl=get_column_letter(col)
            c=ws.cell(total_row,col,f"=SUM({cl}{fr}:{cl}{lr})")
            dc_bg,dc_txt=D_COLORS.get(dasturlar[di]["num"],("DBEAFE","1E40AF"))
            c.fill=fill(dc_bg); c.font=font(bold=True,color=dc_txt,size=10)
            c.alignment=al("center"); c.border=border(); c.number_format="0"
        sum_col=3+hf*(n_d+1)+n_d
        cl=get_column_letter(sum_col)
        c=ws.cell(total_row,sum_col,f"=SUM({cl}{fr}:{cl}{lr})")
        c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=11)
        c.alignment=al("center"); c.border=thick_border(); c.number_format="0"

    ws.freeze_panes="C4"

# ── UMUMIY JADVAL (manager × dastur summary) ────────────────────────────
def write_summary(wb, dasturlar, all_managers):
    ws = wb.create_sheet("UMUMIY JADVAL", 1)
    ws.sheet_view.showGridLines = False
    ws.tab_color = "059669"

    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"]==name), None)
            if m: t += sum(m["weeks"])
        return t
    sorted_mgr = sorted(all_managers, key=grand_total, reverse=True)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    for di in range(len(dasturlar)):
        ws.column_dimensions[get_column_letter(3+di*2)].width = 8
        ws.column_dimensions[get_column_letter(4+di*2)].width = 7

    # ROW 1: Title
    ws.row_dimensions[1].height = 28
    total_end = 2 + len(dasturlar)*2
    c=ws.cell(1,1,"MENEJERLAR UMUMIY JADVALI — Barcha dasturlar (18–24)")
    c.fill=fill(H1); c.font=font(bold=True,size=13,color=H1_TXT); c.alignment=al("left")
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_end)
    for col in range(1,total_end+1): ws.cell(1,col).fill=fill(H1)

    # ROW 2: Dastur headers
    ws.row_dimensions[2].height = 20
    c=ws.cell(2,1,"Menejer"); c.fill=fill(H2); c.font=font(bold=True,color=H2_TXT); c.alignment=al("left"); c.border=border()
    c=ws.cell(2,2,"GRAND\nTOTAL"); c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=9); c.alignment=al("center",wrap=True); c.border=border()
    ws.merge_cells(start_row=2,start_column=2,end_row=3,end_column=2)
    for di,d in enumerate(dasturlar):
        col1=3+di*2
        dc_bg,dc_txt=D_COLORS.get(d["num"],("DBEAFE","1E40AF"))
        c=ws.cell(2,col1,d["name"])
        c.fill=fill(dc_bg); c.font=font(bold=True,color=dc_txt,size=10); c.alignment=al("center"); c.border=border()
        ws.cell(2,col1+1).fill=fill(dc_bg); ws.cell(2,col1+1).border=border()
        ws.merge_cells(start_row=2,start_column=col1,end_row=2,end_column=col1+1)

    # ROW 3: Sub-headers
    ws.row_dimensions[3].height = 18
    ws.cell(3,1).fill=fill(H2); ws.cell(3,1).border=border()
    for di,d in enumerate(dasturlar):
        col1=3+di*2
        dc_bg,dc_txt=D_COLORS.get(d["num"],("DBEAFE","1E40AF"))
        c=ws.cell(3,col1,"Jami"); c.fill=fill(dc_bg); c.font=font(bold=True,color=dc_txt,size=9); c.alignment=al("center"); c.border=border()
        c=ws.cell(3,col1+1,"O'rt."); c.fill=fill(H3); c.font=font(bold=False,color=H3_TXT,size=9,italic=True); c.alignment=al("center"); c.border=border()

    # Data rows
    first=4
    for ri,name in enumerate(sorted_mgr):
        row=first+ri; ws.row_dimensions[row].height=17
        bg=BG if ri%2==0 else ROW_ALT
        gt=grand_total(name)

        c=ws.cell(row,1,name); c.fill=fill(bg)
        c.font=font(bold=gt>0,color=TEXT if gt>0 else MUTED); c.alignment=al("left"); c.border=border()

        # Grand total formula
        cols=[get_column_letter(3+di*2) for di in range(len(dasturlar))]
        gc=ws.cell(row,2,"="+"+".join(f"{cl}{row}" for cl in cols))
        gc.fill=fill(TOTAL_BG if gt>0 else bg); gc.font=font(bold=True,color=TOTAL_TXT if gt>0 else MUTED,size=11)
        gc.alignment=al("center"); gc.border=thick_border(); gc.number_format="0"

        for di,d in enumerate(dasturlar):
            col1=3+di*2
            m=next((x for x in d["managers"] if x["name"]==name),None)
            total=sum(m["weeks"]) if m else 0
            n_wks_d=len(d["weeks"])
            avg=total/n_wks_d if n_wks_d>0 else 0
            dc_bg,dc_txt=D_COLORS.get(d["num"],("DBEAFE","1E40AF"))

            c=ws.cell(row,col1,total if total else None)
            c.fill=fill(dc_bg if total>0 else bg)
            c.font=font(bold=total>0,color=dc_txt if total>0 else MUTED,size=11 if total>0 else 9)
            c.alignment=al("center"); c.border=border(); c.number_format="0"

            c=ws.cell(row,col1+1,round(avg,1) if avg else None)
            c.fill=fill(bg); c.font=font(color=GOOD if avg>=3 else (H3_TXT if avg>0 else MUTED),size=9)
            c.alignment=al("center"); c.border=border(); c.number_format="0.0"

    # Jami qatori
    total_r=first+len(sorted_mgr); ws.row_dimensions[total_r].height=22
    fr=first; lr=total_r-1
    c=ws.cell(total_r,1,"JAMI"); c.fill=fill(TOTAL_BG); c.font=font(bold=True,color=TOTAL_TXT,size=11); c.alignment=al("left"); c.border=border()
    c=ws.cell(total_r,2,f"=SUM(B{fr}:B{lr})"); c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=12)
    c.alignment=al("center"); c.border=thick_border(); c.number_format="0"
    for di,d in enumerate(dasturlar):
        col1=3+di*2; cl=get_column_letter(col1)
        dc_bg,dc_txt=D_COLORS.get(d["num"],("DBEAFE","1E40AF"))
        c=ws.cell(total_r,col1,f"=SUM({cl}{fr}:{cl}{lr})")
        c.fill=fill(dc_bg); c.font=font(bold=True,color=dc_txt,size=11)
        c.alignment=al("center"); c.border=border(); c.number_format="0"
        c=ws.cell(total_r,col1+1); c.fill=fill(H3); c.border=border()

    ws.freeze_panes="C4"

# ── HAR BIR DASTUR SHEET ─────────────────────────────────────────────────
def write_dastur(wb, d):
    ws = wb.create_sheet(d["name"])
    ws.sheet_view.showGridLines = False
    dc_bg,dc_txt = D_COLORS.get(d["num"],("DBEAFE","1E40AF"))
    ws.tab_color = dc_bg

    mgrs   = sorted(d["managers"], key=lambda m: -m["total"])
    weeks  = d["weeks"]
    n_wk   = len(weeks)
    extra  = 3 + n_wk  # бошқа col start

    # Kengliklari
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 9
    for w in range(n_wk): ws.column_dimensions[get_column_letter(3+w)].width = 6
    ws.column_dimensions[get_column_letter(extra)].width   = 9   # бошқа
    ws.column_dimensions[get_column_letter(extra+1)].width = 8   # otkazlar
    ws.column_dimensions[get_column_letter(extra+2)].width = 8   # o'rtacha

    # ROW 1: Dastur title
    ws.row_dimensions[1].height = 28
    total_end = extra+2
    c=ws.cell(1,1,d["name"]+" — Haftalik sotuv jadvali")
    c.fill=fill(H1); c.font=font(bold=True,size=13,color=H1_TXT); c.alignment=al("left")
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=total_end)
    for col in range(1,total_end+1): ws.cell(1,col).fill=fill(H1)

    # ROW 2: Sanalar (start date)
    ws.row_dimensions[2].height = 15
    ws.cell(2,1,"Boshlanish").fill=fill(H3); ws.cell(2,1).font=font(size=8,color=MUTED,italic=True); ws.cell(2,1).border=border()
    ws.cell(2,2).fill=fill(H3); ws.cell(2,2).border=border()
    for w,wk in enumerate(weeks):
        s=wk["start"][5:].replace("-",".") if wk["start"] else ""
        c=ws.cell(2,3+w,s); c.fill=fill(H3); c.font=font(size=8,color=MUTED,italic=True); c.alignment=al("center"); c.border=border()
    for col in range(extra,extra+3):
        ws.cell(2,col).fill=fill(H3); ws.cell(2,col).border=border()

    # ROW 3: Sanalar (end date)
    ws.row_dimensions[3].height = 15
    ws.cell(3,1,"Tugash").fill=fill(H3); ws.cell(3,1).font=font(size=8,color=MUTED,italic=True); ws.cell(3,1).border=border()
    ws.cell(3,2).fill=fill(H3); ws.cell(3,2).border=border()
    for w,wk in enumerate(weeks):
        s=wk["end"][5:].replace("-",".") if wk["end"] else ""
        c=ws.cell(3,3+w,s); c.fill=fill(H3); c.font=font(size=8,color=MUTED,italic=True); c.alignment=al("center"); c.border=border()
    for col in range(extra,extra+3):
        ws.cell(3,col).fill=fill(H3); ws.cell(3,col).border=border()

    # ROW 4: Header
    ws.row_dimensions[4].height = 22
    hdrs=["Menejer","Jami"]+[f"{w+1}-hf" for w in range(n_wk)]+["Boshqa","Otkazlar","O'rtacha"]
    for col,h in enumerate(hdrs,1):
        is_jami = col==2
        c=ws.cell(4,col,h)
        c.fill=fill(H1 if is_jami else (dc_bg if col==1 or col>2+n_wk else H2))
        c.font=font(bold=True, color=H1_TXT if is_jami else (dc_txt if col==1 or col>2+n_wk else H2_TXT), size=9)
        c.alignment=al("center" if col>1 else "left"); c.border=border()

    # Data rows
    first=5
    for ri,m in enumerate(mgrs):
        row=first+ri; ws.row_dimensions[row].height=17
        bg=BG if ri%2==0 else ROW_ALT
        is_zero=m["total"]==0
        avg=m["total"]/n_wk if n_wk>0 else 0

        # Menejer nomi
        c=ws.cell(row,1,m["name"]); c.fill=fill(bg)
        c.font=font(bold=not is_zero, color=TEXT if not is_zero else MUTED); c.alignment=al("left"); c.border=border()

        # Jami (formula)
        wk_first=get_column_letter(3); wk_last=get_column_letter(2+n_wk)
        c=ws.cell(row,2,f"=SUM({wk_first}{row}:{wk_last}{row})")
        c.fill=fill(dc_bg if not is_zero else bg)
        c.font=font(bold=True,color=dc_txt if not is_zero else MUTED,size=11)
        c.alignment=al("center"); c.border=thick_border(); c.number_format="0"

        # Hafta values
        mx=max(m["weeks"]) if m["weeks"] else 1
        for w,v in enumerate(m["weeks"]):
            c=ws.cell(row,3+w,v if v else None)
            if v==0:
                c.fill=fill(bg); c.font=font(color="CBD5E1",size=9)
            elif v==mx and v>1:
                c.fill=fill("1D4ED8"); c.font=font(bold=True,color="FFFFFF",size=10)
            elif v>=mx*0.6:
                c.fill=fill("3B82F6"); c.font=font(bold=True,color="FFFFFF",size=10)
            elif v>=mx*0.3:
                c.fill=fill(H2); c.font=font(bold=True,color=H2_TXT,size=10)
            else:
                c.fill=fill(H3); c.font=font(color=H3_TXT,size=10)
            c.alignment=al("center"); c.border=border(); c.number_format="0"

        # Бошқа
        c=ws.cell(row,extra,m["boshqa"] if m["boshqa"] else None)
        c.fill=fill(bg); c.font=font(color="0F766E" if m["boshqa"] else MUTED); c.alignment=al("center"); c.border=border()

        # Otkazlar
        c=ws.cell(row,extra+1,m["otkazlar"] if m["otkazlar"] else None)
        c.fill=fill(bg); c.font=font(color="DC2626" if m["otkazlar"] else MUTED); c.alignment=al("center"); c.border=border()

        # O'rtacha (formula)
        c=ws.cell(row,extra+2,f'=IFERROR(B{row}/{n_wk},"")')
        c.fill=fill(bg); c.font=font(color=GOOD if avg>=3 else H3_TXT,bold=avg>=3); c.alignment=al("center"); c.border=border(); c.number_format="0.0"

    # Jami qatori
    total_r=first+len(mgrs); ws.row_dimensions[total_r].height=22
    fr=first; lr=total_r-1
    c=ws.cell(total_r,1,"JAMI"); c.fill=fill(TOTAL_BG); c.font=font(bold=True,color=TOTAL_TXT,size=11); c.alignment=al("left"); c.border=border()
    c=ws.cell(total_r,2,f"=SUM(B{fr}:B{lr})"); c.fill=fill(H1); c.font=font(bold=True,color=H1_TXT,size=12)
    c.alignment=al("center"); c.border=thick_border(); c.number_format="0"
    for w in range(n_wk):
        cl=get_column_letter(3+w)
        c=ws.cell(total_r,3+w,f"=SUM({cl}{fr}:{cl}{lr})")
        c.fill=fill(H2); c.font=font(bold=True,color=H2_TXT)
        c.alignment=al("center"); c.border=border(); c.number_format="0"
    for off,clr in [(0,"0F766E"),(1,"DC2626")]:
        cl=get_column_letter(extra+off)
        c=ws.cell(total_r,extra+off,f"=SUM({cl}{fr}:{cl}{lr})")
        c.fill=fill(TOTAL_BG); c.font=font(bold=True,color=clr); c.alignment=al("center"); c.border=border(); c.number_format="0"
    c=ws.cell(total_r,extra+2,f'=IFERROR(B{total_r}/{n_wk},"")')
    c.fill=fill(TOTAL_BG); c.font=font(bold=True,color=GOOD); c.alignment=al("center"); c.border=border(); c.number_format="0.0"

    ws.freeze_panes="C5"

# ── HAFTALIK SHEET: Menejer × Hafta × Dastur (Haftalik.jpg ko'rinishi) ──
def write_haftalik(wb, dasturlar, all_managers):
    ws = wb.create_sheet("HAFTALIK")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "1E3A8A"

    NAV_BG  = "1E3A8A"
    NAV_TXT = "FFFFFF"
    HF_BG   = "1D4ED8"
    HF_TXT  = "FFFFFF"
    SUB_BG  = "DBEAFE"
    SUB_TXT = "1E3A8A"

    n_d         = len(dasturlar)
    max_wk      = max(len(d["weeks"]) for d in dasturlar)
    total_weeks = sum(len(d["weeks"]) for d in dasturlar)  # barcha hafta soni (o'rtacha uchun bo'luvchi)

    # Ustun rejasi: A=Menejer | B=O'rtacha | C=Jami | D+ = data (hafta×dastur)
    # Data ustun col = 4 + hf*n_d + di
    DATA_START  = 4
    last_data_col = get_column_letter(DATA_START - 1 + max_wk * n_d)
    total_cols  = 3 + max_wk * n_d

    # Ustun kengliklari
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 8   # O'rtacha
    ws.column_dimensions["C"].width = 7   # Jami
    for hf in range(max_wk):
        for di in range(n_d):
            ws.column_dimensions[get_column_letter(DATA_START + hf*n_d + di)].width = 6

    # ROW 1: Sarlavha
    ws.row_dimensions[1].height = 28
    c = ws.cell(1, 1, "HAFTALIK SOTUV JADVALI  —  Menejer × Hafta × Dastur")
    c.fill = fill(NAV_BG); c.font = font(bold=True, size=13, color=NAV_TXT)
    c.alignment = al("left"); c.border = border()
    for col in range(2, total_cols+1):
        ws.cell(1, col).fill = fill(NAV_BG)
        ws.cell(1, col).border = border()
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # ROW 2-3: Menejer / O'rtacha / Jami (merged 2 rows) + hafta merged headerlar
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 18

    for col_idx, label in [(1,"Menejer"), (2,"O'rtacha"), (3,"Jami")]:
        c = ws.cell(2, col_idx, label)
        c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10 if col_idx==1 else 9)
        c.alignment = al("left" if col_idx==1 else "center"); c.border = border()
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
        ws.cell(3, col_idx).fill = fill(NAV_BG); ws.cell(3, col_idx).border = border()

    for hf in range(max_wk):
        first_col = DATA_START + hf * n_d
        last_col  = first_col + n_d - 1
        c = ws.cell(2, first_col, f"{hf+1}-hafta")
        c.fill = fill(HF_BG); c.font = font(bold=True, color=HF_TXT, size=11)
        c.alignment = al("center"); c.border = border()
        for col in range(first_col+1, last_col+1):
            ws.cell(2, col).fill = fill(HF_BG); ws.cell(2, col).border = border()
        ws.merge_cells(start_row=2, start_column=first_col, end_row=2, end_column=last_col)
        # Dastur sub-headerlar (row 3)
        for di, d in enumerate(dasturlar):
            col = DATA_START + hf*n_d + di
            c = ws.cell(3, col, f"{d['num']}-dastur")
            c.fill = fill(SUB_BG); c.font = font(bold=False, color=SUB_TXT, size=8)
            c.alignment = al("center"); c.border = border()

    # Menejerlani grand total bo'yicha saralash
    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"]==name), None)
            if m: t += sum(m["weeks"])
        return t
    sorted_mgr = sorted(all_managers, key=grand_total, reverse=True)

    # ROW 4+: Menejer datalari
    first_data = 4
    for ri, name in enumerate(sorted_mgr):
        row = first_data + ri
        ws.row_dimensions[row].height = 17
        bg  = BG if ri%2==0 else ROW_ALT
        gt  = grand_total(name)
        avg = gt / total_weeks if total_weeks > 0 else 0

        # A: Menejer nomi (dark navy)
        c = ws.cell(row, 1, name)
        c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10)
        c.alignment = al("left"); c.border = border()

        # B: O'rtacha = Jami / total_weeks (barcha dastur haftalar bo'yicha)
        c = ws.cell(row, 2, f"=IFERROR(C{row}/{total_weeks},\"\")")
        c.fill = fill(HF_BG if avg >= 1 else NAV_BG)
        c.font = font(bold=avg >= 1, color=HF_TXT, size=10)
        c.alignment = al("center"); c.border = border(); c.number_format = "0.0"

        # C: Jami (sum of all data cells)
        c = ws.cell(row, 3, f"=SUM(D{row}:{last_data_col}{row})")
        c.fill = fill(HF_BG if gt else NAV_BG)
        c.font = font(bold=True, color=HF_TXT, size=11)
        c.alignment = al("center"); c.border = border(); c.number_format = "0"

        # D+: Har bir hafta × dastur
        for hf in range(max_wk):
            for di, d in enumerate(dasturlar):
                col = DATA_START + hf*n_d + di
                m   = next((x for x in d["managers"] if x["name"]==name), None)
                v   = m["weeks"][hf] if m and hf < len(m["weeks"]) else 0

                c = ws.cell(row, col)
                if v:
                    c.value = v
                    c.fill  = fill(SUB_BG)
                    c.font  = font(bold=True, color=SUB_TXT, size=10)
                else:
                    c.value = None
                    c.fill  = fill(bg)
                    c.font  = font(color="CBD5E1", size=9)
                c.alignment = al("center"); c.border = border(); c.number_format = "0"

    # JAMI qatori
    total_r = first_data + len(sorted_mgr)
    ws.row_dimensions[total_r].height = 22
    fr = first_data; lr = total_r - 1

    c = ws.cell(total_r, 1, "JAMI")
    c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=11)
    c.alignment = al("left"); c.border = border()

    c = ws.cell(total_r, 2, f"=IFERROR(C{total_r}/{total_weeks},\"\")")
    c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=10)
    c.alignment = al("center"); c.border = border(); c.number_format = "0.0"

    c = ws.cell(total_r, 3, f"=SUM(C{fr}:C{lr})")
    c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=12)
    c.alignment = al("center"); c.border = thick_border(); c.number_format = "0"

    for hf in range(max_wk):
        for di in range(n_d):
            col = DATA_START + hf*n_d + di
            cl  = get_column_letter(col)
            c = ws.cell(total_r, col, f"=SUM({cl}{fr}:{cl}{lr})")
            c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=10)
            c.alignment = al("center"); c.border = border(); c.number_format = "0"

    ws.freeze_panes = "D4"

    # ── QO'SHIMCHA JADVAL: 8 HAFTAGACHA JAMI SOTUV ──────────────────────
    WK_LIMIT = 8
    tbl2_title_row = total_r + 3   # 2 qator oraliq
    tbl2_hdr_row   = tbl2_title_row + 1
    tbl2_first     = tbl2_hdr_row + 1
    # Ustunlar: A=Menejer | B=Jami(1-8hf) | C..I = 7 dastur
    tbl2_end_col   = 2 + n_d      # 2 + 7 = 9

    # 2.1 — Sarlavha
    ws.row_dimensions[tbl2_title_row].height = 26
    c = ws.cell(tbl2_title_row, 1,
                f"BIRINCHI {WK_LIMIT} HAFTA SOTUV  —  Menejer × Dastur")
    c.fill = fill(NAV_BG); c.font = font(bold=True, size=12, color=NAV_TXT)
    c.alignment = al("left"); c.border = border()
    for col in range(2, tbl2_end_col + 1):
        ws.cell(tbl2_title_row, col).fill   = fill(NAV_BG)
        ws.cell(tbl2_title_row, col).border = border()
    ws.merge_cells(start_row=tbl2_title_row, start_column=1,
                   end_row=tbl2_title_row,   end_column=tbl2_end_col)

    # 2.2 — Header qatori
    ws.row_dimensions[tbl2_hdr_row].height = 20
    c = ws.cell(tbl2_hdr_row, 1, "Menejer")
    c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10)
    c.alignment = al("left"); c.border = border()

    c = ws.cell(tbl2_hdr_row, 2, f"Jami\n1–{WK_LIMIT}hf")
    c.fill = fill(HF_BG); c.font = font(bold=True, color=HF_TXT, size=9)
    c.alignment = al("center", wrap=True); c.border = border()

    for di, d in enumerate(dasturlar):
        col = 3 + di
        wk_n = min(WK_LIMIT, len(d["weeks"]))
        dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))
        c = ws.cell(tbl2_hdr_row, col, f"{d['num']}-D\n1–{wk_n}hf")
        c.fill = fill(dc_bg); c.font = font(bold=True, color=dc_txt, size=8)
        c.alignment = al("center", wrap=True); c.border = border()

    # 2.3 — Ma'lumot qatorlari
    def sum8(name):
        total = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"] == name), None)
            if m:
                total += sum(m["weeks"][:min(WK_LIMIT, len(m["weeks"]))])
        return total

    sorted_mgr8 = sorted(all_managers, key=sum8, reverse=True)

    for ri, name in enumerate(sorted_mgr8):
        row = tbl2_first + ri
        ws.row_dimensions[row].height = 17
        bg  = BG if ri % 2 == 0 else ROW_ALT
        t8  = sum8(name)

        c = ws.cell(row, 1, name)
        c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10)
        c.alignment = al("left"); c.border = border()

        c = ws.cell(row, 2, t8 if t8 else None)
        c.fill = fill(HF_BG if t8 else bg)
        c.font = font(bold=True, color=HF_TXT if t8 else MUTED, size=11)
        c.alignment = al("center"); c.border = border(); c.number_format = "0"

        for di, d in enumerate(dasturlar):
            col = 3 + di
            m   = next((x for x in d["managers"] if x["name"] == name), None)
            v   = sum(m["weeks"][:min(WK_LIMIT, len(m["weeks"]))]) if m else 0
            dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))

            c = ws.cell(row, col)
            if v:
                c.value = v
                c.fill  = fill(dc_bg)
                c.font  = font(bold=True, color=dc_txt, size=10)
            else:
                c.value = None
                c.fill  = fill(bg)
                c.font  = font(color="CBD5E1", size=9)
            c.alignment = al("center"); c.border = border(); c.number_format = "0"

    # 2.4 — Jami qatori
    tbl2_total = tbl2_first + len(sorted_mgr8)
    ws.row_dimensions[tbl2_total].height = 22
    fr2 = tbl2_first; lr2 = tbl2_total - 1

    c = ws.cell(tbl2_total, 1, "JAMI")
    c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=11)
    c.alignment = al("left"); c.border = border()

    cl2 = get_column_letter(2)
    c = ws.cell(tbl2_total, 2, f"=SUM({cl2}{fr2}:{cl2}{lr2})")
    c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=12)
    c.alignment = al("center"); c.border = thick_border(); c.number_format = "0"

    for di in range(n_d):
        col = 3 + di
        cl  = get_column_letter(col)
        c = ws.cell(tbl2_total, col, f"=SUM({cl}{fr2}:{cl}{lr2})")
        c.fill = fill(H1); c.font = font(bold=True, color=H1_TXT, size=10)
        c.alignment = al("center"); c.border = border(); c.number_format = "0"

# ── MAIN ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("CSV yuklanmoqda...")
    dasturlar = fetch()
    for d in dasturlar:
        act=len([m for m in d["managers"] if m["total"]>0])
        print(f"  {d['name']}: {len(d['weeks'])} hafta, {act} faol menejer")

    all_managers = sorted({m["name"] for d in dasturlar for m in d["managers"]})
    print(f"Jami: {len(all_managers)} menejer, {len(dasturlar)} dastur")

    wb = Workbook()
    wb.remove(wb.active)

    write_pivot(wb, dasturlar, all_managers)
    write_summary(wb, dasturlar, all_managers)
    write_haftalik(wb, dasturlar, all_managers)
    for d in dasturlar:
        write_dastur(wb, d)

    wb.save(OUTPUT)
    print(f"\n✓ Saqlandi: {OUTPUT}")
