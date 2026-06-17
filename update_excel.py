"""
Mavjud Excel fayliga yangi menejerlarni qo'shadi.
- HAFTALIK sheet: user o'zgartirmalarini saqlab, yangi qatorlar kiritadi
- Qolgan sheetlar: yangi data bilan qayta yaratiladi
"""
import re, sys
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

# generate_excel.py dan barcha funksiyalar va konstantalarni import qilamiz
sys.path.insert(0, "/Users/mac/Work Milliard/Dastur 24")
from generate_excel import (
    fetch, fill, font, border, thick_border, al,
    write_pivot, write_summary, write_dastur,
    BG, ROW_ALT, H1, H1_TXT, H2, H2_TXT, H3, H3_TXT,
    TOTAL_BG, TOTAL_TXT, TEXT, MUTED, GOOD, BORDER_C, D_COLORS,
    OUTPUT
)

EXCEL_PATH = OUTPUT
NAV_BG = "1E3A8A"; NAV_TXT = "FFFFFF"
HF_BG  = "1D4ED8"; HF_TXT  = "FFFFFF"
SUB_BG = "DBEAFE"; SUB_TXT = "1E3A8A"


def find_jami_rows(ws):
    """HAFTALIK sheetdagi barcha 'JAMI' qatorlarini topadi"""
    rows = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        if row[0].value == "JAMI":
            rows.append(row[0].row)
    return rows


def get_sheet_managers(ws, first_data_row=4):
    """first_jami gacha bo'lgan col A dan menejer nomlarini oladi"""
    jami_rows = find_jami_rows(ws)
    if not jami_rows:
        return set(), None
    first_jami = jami_rows[0]
    mgrs = set()
    for r in range(first_data_row, first_jami):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str):
            mgrs.add(v)
    return mgrs, first_jami


def update_formula_row_refs(formula, shift_from_row, shift_by):
    """
    Formuladagi row raqamlarini yangilaydi:
    shift_from_row dan katta yoki teng bo'lgan har bir row raqamiga shift_by qo'shadi.
    Masalan: =SUM(B25:B37) + shift_from=21, shift_by=2 → =SUM(B27:B39)
    """
    def replacer(m):
        col = m.group(1)
        row_num = int(m.group(2))
        if row_num >= shift_from_row:
            return f"{col}{row_num + shift_by}"
        return m.group(0)
    return re.sub(r'([A-Z]+)(\d+)', replacer, formula)


def rebuild_main_jami_row(ws, jami_row, first_data, new_last_data,
                           n_d, max_wk, total_weeks, DATA_START):
    """Asosiy JAMI qatorining formulalarini qaytadan yozadi"""
    last_data_col = get_column_letter(DATA_START - 1 + max_wk * n_d)

    # Col A: "JAMI"
    c = ws.cell(jami_row, 1)
    if c.value != "JAMI":
        c.value = "JAMI"

    # Col B: O'rtacha
    c = ws.cell(jami_row, 2, f'=IFERROR(C{jami_row}/{total_weeks},"")')
    c.number_format = "0.0"

    # Col C: Jami
    c = ws.cell(jami_row, 3, f"=SUM(C{first_data}:C{new_last_data})")
    c.number_format = "0"

    # D+ data cols
    for hf in range(max_wk):
        for di in range(n_d):
            col = DATA_START + hf * n_d + di
            cl = get_column_letter(col)
            c = ws.cell(jami_row, col, f"=SUM({cl}{first_data}:{cl}{new_last_data})")
            c.number_format = "0"


def fix_second_jami_row(ws, second_jami_row, shift_from, shift_by):
    """8-hafta bo'limi JAMI formulalarini yangilaydi"""
    for cell in ws[second_jami_row]:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
            cell.value = update_formula_row_refs(cell.value, shift_from, shift_by)


def add_to_haftalik(ws, dasturlar, new_managers, all_managers):
    """HAFTALIK sheetga yangi menejer qatorlarini kiritadi"""
    if not new_managers:
        return

    n_d         = len(dasturlar)
    max_wk      = max(len(d["weeks"]) for d in dasturlar)
    total_weeks = sum(len(d["weeks"]) for d in dasturlar)
    DATA_START  = 4
    FIRST_DATA  = 4
    last_data_col = get_column_letter(DATA_START - 1 + max_wk * n_d)

    jami_rows = find_jami_rows(ws)
    if not jami_rows:
        print("  HAFTALIK: JAMI qatori topilmadi, o'tkazib yuborildi.")
        return

    first_jami = jami_rows[0]
    old_last_data = first_jami - 1
    N = len(new_managers)

    print(f"  HAFTALIK: {first_jami}-qator oldiga {N} ta yangi qator kiritilmoqda...")
    ws.insert_rows(first_jami, amount=N)
    # Endi: yangi qatorlar = first_jami .. first_jami+N-1
    # Asosiy JAMI ko'chdi = first_jami + N
    # 8-hafta JAMI (agar mavjud) = jami_rows[1] + N

    new_last_data = first_jami + N - 1
    new_jami_row  = first_jami + N

    # --- Yangi menejer qatorlarini yozamiz ---
    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"] == name), None)
            if m: t += sum(m["weeks"])
        return t

    # Avvalgi alternating tartibini davom ettiramiz
    last_existing_ri = old_last_data - FIRST_DATA  # 0-indexed
    for ri, name in enumerate(new_managers):
        row = first_jami + ri
        ws.row_dimensions[row].height = 17
        bg = BG if (last_existing_ri + 1 + ri) % 2 == 0 else ROW_ALT
        gt = grand_total(name)

        # A: Menejer
        c = ws.cell(row, 1, name)
        c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10)
        c.alignment = al("left"); c.border = border()

        # B: O'rtacha
        c = ws.cell(row, 2, f'=IFERROR(C{row}/{total_weeks},"")')
        c.fill = fill(HF_BG if gt else NAV_BG)
        c.font = font(bold=gt > 0, color=HF_TXT, size=10)
        c.alignment = al("center"); c.border = border(); c.number_format = "0.0"

        # C: Jami
        c = ws.cell(row, 3, f"=SUM(D{row}:{last_data_col}{row})")
        c.fill = fill(HF_BG if gt else NAV_BG)
        c.font = font(bold=True, color=HF_TXT, size=11)
        c.alignment = al("center"); c.border = border(); c.number_format = "0"

        # D+: hafta × dastur
        for hf in range(max_wk):
            for di, d in enumerate(dasturlar):
                col = DATA_START + hf * n_d + di
                m = next((x for x in d["managers"] if x["name"] == name), None)
                v = m["weeks"][hf] if m and hf < len(m["weeks"]) else 0

                c = ws.cell(row, col)
                if v:
                    c.value = v
                    c.fill  = fill(SUB_BG)
                    c.font  = font(bold=True, color=SUB_TXT, size=10)
                else:
                    c.value = None
                    c.fill  = fill(bg)
                    c.font  = font(color="CBD5E1", size=9)
                c.alignment = al("center"); c.border = border(); c.number_format = "0"

    # --- Asosiy JAMI qatori formulalarini yangilaymiz ---
    rebuild_main_jami_row(ws, new_jami_row, FIRST_DATA, new_last_data,
                          n_d, max_wk, total_weeks, DATA_START)

    # --- 8-hafta bo'limi JAMI formulalarini yangilaymiz ---
    if len(jami_rows) > 1:
        old_second_jami = jami_rows[1]
        new_second_jami = old_second_jami + N
        fix_second_jami_row(ws, new_second_jami, shift_from=first_jami, shift_by=N)
        print(f"  HAFTALIK: 8-hafta JAMI ({new_second_jami}-qator) formulalari yangilandi.")

    print(f"  HAFTALIK: {N} ta yangi menejer qo'shildi.")


def add_to_umumiy(ws, dasturlar, new_managers):
    """UMUMIY JADVAL sheetga yangi menejer qatorlarini kiritadi"""
    if not new_managers:
        return

    FIRST_DATA = 4
    jami_rows = find_jami_rows(ws)
    if not jami_rows:
        return
    first_jami = jami_rows[0]
    old_last_data = first_jami - 1
    N = len(new_managers)

    ws.insert_rows(first_jami, amount=N)
    new_jami_row  = first_jami + N
    new_last_data = first_jami + N - 1
    last_existing_ri = old_last_data - FIRST_DATA

    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"] == name), None)
            if m: t += sum(m["weeks"])
        return t

    for ri, name in enumerate(new_managers):
        row = first_jami + ri
        ws.row_dimensions[row].height = 17
        bg = BG if (last_existing_ri + 1 + ri) % 2 == 0 else ROW_ALT
        gt = grand_total(name)

        # A: Menejer
        c = ws.cell(row, 1, name)
        c.fill = fill(bg); c.font = font(bold=gt > 0, color=TEXT if gt > 0 else MUTED)
        c.alignment = al("left"); c.border = border()

        # B: Grand Total formula
        cols = [get_column_letter(3 + di * 2) for di in range(len(dasturlar))]
        gc = ws.cell(row, 2, "=" + "+".join(f"{cl}{row}" for cl in cols))
        gc.fill = fill(TOTAL_BG if gt > 0 else bg)
        gc.font = font(bold=True, color=TOTAL_TXT if gt > 0 else MUTED, size=11)
        gc.alignment = al("center"); gc.border = thick_border(); gc.number_format = "0"

        # Per-dastur: jami + o'rtacha
        for di, d in enumerate(dasturlar):
            col1 = 3 + di * 2
            m = next((x for x in d["managers"] if x["name"] == name), None)
            total = sum(m["weeks"]) if m else 0
            n_wks_d = len(d["weeks"])
            avg = total / n_wks_d if n_wks_d > 0 else 0
            dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))

            c = ws.cell(row, col1, total if total else None)
            c.fill = fill(dc_bg if total > 0 else bg)
            c.font = font(bold=total > 0, color=dc_txt if total > 0 else MUTED, size=11 if total > 0 else 9)
            c.alignment = al("center"); c.border = border(); c.number_format = "0"

            c = ws.cell(row, col1 + 1, round(avg, 1) if avg else None)
            c.fill = fill(bg); c.font = font(color=GOOD if avg >= 3 else (H3_TXT if avg > 0 else MUTED), size=9)
            c.alignment = al("center"); c.border = border(); c.number_format = "0.0"

    # JAMI formulalarini yangilaymiz
    for cell in ws[new_jami_row]:
        if cell.value and isinstance(cell.value, str) and "SUM" in cell.value:
            col_l = get_column_letter(cell.column)
            cell.value = f"=SUM({col_l}{FIRST_DATA}:{col_l}{new_last_data})"

    print(f"  UMUMIY JADVAL: {N} ta yangi menejer qo'shildi.")


def add_to_pivot(ws, dasturlar, new_managers):
    """HAFTA × DASTUR × MENEJER sheetga yangi qatorlar kiritadi"""
    if not new_managers:
        return

    FIRST_DATA = 4
    n_d    = len(dasturlar)
    max_wk = max(len(d["weeks"]) for d in dasturlar)

    jami_rows = find_jami_rows(ws)
    if not jami_rows:
        return
    first_jami = jami_rows[0]
    old_last_data = first_jami - 1
    N = len(new_managers)

    ws.insert_rows(first_jami, amount=N)
    new_jami_row  = first_jami + N
    new_last_data = first_jami + N - 1
    last_existing_ri = old_last_data - FIRST_DATA

    def grand_total(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"] == name), None)
            if m: t += sum(m["weeks"])
        return t

    for ri, name in enumerate(new_managers):
        row = first_jami + ri
        ws.row_dimensions[row].height = 17
        bg = BG if (last_existing_ri + 1 + ri) % 2 == 0 else ROW_ALT
        gt = grand_total(name)

        # A: Menejer
        c = ws.cell(row, 1, name)
        c.fill = fill(bg); c.font = font(bold=gt > 0, color=TEXT if gt > 0 else MUTED, size=10)
        c.alignment = al("left"); c.border = border()

        # B: Grand Total
        sum_cols = [get_column_letter(3 + hf * (n_d + 1) + n_d) for hf in range(max_wk)]
        formula = "=" + "+".join(f"{cl}{row}" for cl in sum_cols)
        gc = ws.cell(row, 2, formula)
        gc.fill = fill(TOTAL_BG if gt > 0 else bg)
        gc.font = font(bold=True, color=TOTAL_TXT if gt > 0 else MUTED, size=11)
        gc.alignment = al("center"); gc.border = thick_border(); gc.number_format = "0"

        # Hafta × Dastur values
        for hf in range(max_wk):
            hafta_vals_cols = []
            for di, d in enumerate(dasturlar):
                col = 3 + hf * (n_d + 1) + di
                m   = next((x for x in d["managers"] if x["name"] == name), None)
                v   = m["weeks"][hf] if m and hf < len(m["weeks"]) else None

                c = ws.cell(row, col)
                if v is None:
                    c.value = None; c.fill = fill("F1F5F9"); c.font = font(color="E2E8F0", size=9)
                elif v == 0:
                    c.value = None; c.fill = fill(bg); c.font = font(color="CBD5E1", size=9)
                else:
                    c.value = v
                    dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))
                    mx = max((x["weeks"][hf] if hf < len(x["weeks"]) else 0) for x in d["managers"]) or 1
                    if v == mx and v > 1:
                        c.fill = fill("1D4ED8"); c.font = font(bold=True, color="FFFFFF", size=10)
                    elif v >= mx * 0.6:
                        c.fill = fill("3B82F6"); c.font = font(bold=True, color="FFFFFF", size=10)
                    elif v >= mx * 0.3:
                        c.fill = fill(dc_bg); c.font = font(bold=True, color=dc_txt, size=10)
                    else:
                        c.fill = fill(H3); c.font = font(color=H3_TXT, size=10)
                c.alignment = al("center"); c.border = border(); c.number_format = "0"
                hafta_vals_cols.append(get_column_letter(col))

            # Hafta jami
            sum_col = 3 + hf * (n_d + 1) + n_d
            rng = "+".join(f"{cl}{row}" for cl in hafta_vals_cols)
            sc = ws.cell(row, sum_col, f"={rng}")
            hf_total = sum(
                (next((x for x in d["managers"] if x["name"] == name), None) or {"weeks": []})
                ["weeks"][hf] if hf < len(
                    (next((x for x in d["managers"] if x["name"] == name), None) or {"weeks": []})["weeks"]
                ) else 0
                for d in dasturlar
            )
            sc.fill = fill(H2 if hf_total > 0 else ROW_ALT)
            sc.font = font(bold=hf_total > 0, color=H2_TXT if hf_total > 0 else MUTED, size=10)
            sc.alignment = al("center"); sc.border = thick_border(); sc.number_format = "0"

    # JAMI formulalarini yangilaymiz
    for cell in ws[new_jami_row]:
        if cell.value and isinstance(cell.value, str) and "SUM" in cell.value:
            col_l = get_column_letter(cell.column)
            cell.value = f"=SUM({col_l}{FIRST_DATA}:{col_l}{new_last_data})"

    print(f"  PIVOT: {N} ta yangi menejer qo'shildi.")


def add_to_dastur(ws, d, new_managers_in_dastur):
    """Per-dastur sheetga yangi menejer qatorlarini kiritadi"""
    if not new_managers_in_dastur:
        return

    n_wk       = len(d["weeks"])
    extra      = 3 + n_wk
    FIRST_DATA = 5
    dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))

    jami_rows = find_jami_rows(ws)
    if not jami_rows:
        return
    first_jami    = jami_rows[0]
    old_last_data = first_jami - 1
    N = len(new_managers_in_dastur)

    ws.insert_rows(first_jami, amount=N)
    new_jami_row  = first_jami + N
    new_last_data = first_jami + N - 1
    last_existing_ri = old_last_data - FIRST_DATA

    for ri, name in enumerate(new_managers_in_dastur):
        row = first_jami + ri
        ws.row_dimensions[row].height = 17
        bg = BG if (last_existing_ri + 1 + ri) % 2 == 0 else ROW_ALT

        m = next((x for x in d["managers"] if x["name"] == name), None)
        is_zero = (not m or m["total"] == 0)
        avg = sum(m["weeks"]) / n_wk if m and n_wk > 0 else 0

        # A: Menejer
        c = ws.cell(row, 1, name)
        c.fill = fill(bg); c.font = font(bold=not is_zero, color=TEXT if not is_zero else MUTED)
        c.alignment = al("left"); c.border = border()

        # B: Jami (formula)
        wk_first = get_column_letter(3); wk_last = get_column_letter(2 + n_wk)
        c = ws.cell(row, 2, f"=SUM({wk_first}{row}:{wk_last}{row})")
        c.fill = fill(dc_bg if not is_zero else bg)
        c.font = font(bold=True, color=dc_txt if not is_zero else MUTED, size=11)
        c.alignment = al("center"); c.border = thick_border(); c.number_format = "0"

        # Hafta values
        mx = max(m["weeks"]) if m and m["weeks"] else 1
        for w in range(n_wk):
            v = m["weeks"][w] if m and w < len(m["weeks"]) else 0
            c = ws.cell(row, 3 + w, v if v else None)
            if v == 0:
                c.fill = fill(bg); c.font = font(color="CBD5E1", size=9)
            elif v == mx and v > 1:
                c.fill = fill("1D4ED8"); c.font = font(bold=True, color="FFFFFF", size=10)
            elif v >= mx * 0.6:
                c.fill = fill("3B82F6"); c.font = font(bold=True, color="FFFFFF", size=10)
            elif v >= mx * 0.3:
                c.fill = fill(H2); c.font = font(bold=True, color=H2_TXT, size=10)
            else:
                c.fill = fill(H3); c.font = font(color=H3_TXT, size=10)
            c.alignment = al("center"); c.border = border(); c.number_format = "0"

        # Boshqa / Otkazlar
        boshqa   = m["boshqa"]   if m else 0
        otkazlar = m["otkazlar"] if m else 0
        c = ws.cell(row, extra, boshqa if boshqa else None)
        c.fill = fill(bg); c.font = font(color="0F766E" if boshqa else MUTED)
        c.alignment = al("center"); c.border = border()

        c = ws.cell(row, extra + 1, otkazlar if otkazlar else None)
        c.fill = fill(bg); c.font = font(color="DC2626" if otkazlar else MUTED)
        c.alignment = al("center"); c.border = border()

        # O'rtacha (formula)
        c = ws.cell(row, extra + 2, f'=IFERROR(B{row}/{n_wk},"")')
        c.fill = fill(bg); c.font = font(color=GOOD if avg >= 3 else H3_TXT, bold=avg >= 3)
        c.alignment = al("center"); c.border = border(); c.number_format = "0.0"

    # JAMI formulalarini yangilaymiz
    for cell in ws[new_jami_row]:
        if cell.value and isinstance(cell.value, str) and "SUM" in cell.value:
            col_l = get_column_letter(cell.column)
            cell.value = f"=SUM({col_l}{FIRST_DATA}:{col_l}{new_last_data})"
        elif cell.value and isinstance(cell.value, str) and "IFERROR" in cell.value:
            cell.value = f'=IFERROR(B{new_jami_row}/{n_wk},"")'

    print(f"  {d['name']}: {N} ta yangi menejer qo'shildi.")


def sync_8hafta_section(ws, dasturlar, all_managers):
    """
    HAFTALIK sheetdagi '8 HAFTAGACHA' jadvalini sinxronlashtiradi.
    Asosiy jadvaldagi barcha menejerlar 8-hafta jadvalida ham bo'lishini ta'minlaydi.
    """
    WK_LIMIT = 8
    n_d = len(dasturlar)

    # 8-hafta bo'limining title qatorini topamiz ("BIRINCHI" so'zi bilan)
    tbl2_title_row = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v and isinstance(v, str) and "BIRINCHI" in v:
            tbl2_title_row = row[0].row
            break

    if tbl2_title_row is None:
        print("  8-hafta bo'limi topilmadi, o'tkazib yuborildi.")
        return

    tbl2_first = tbl2_title_row + 2   # title + header = 2 qator

    # Ikkinchi JAMI qatorini topamiz
    jami_rows = find_jami_rows(ws)
    if len(jami_rows) < 2:
        print("  8-hafta JAMI topilmadi, o'tkazib yuborildi.")
        return
    second_jami = jami_rows[1]

    # 8-hafta bo'limidagi mavjud menejerlar
    existing_8hf = set()
    for r in range(tbl2_first, second_jami):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str):
            existing_8hf.add(v)

    missing = [m for m in all_managers if m not in existing_8hf]
    if not missing:
        print("  8-hafta bo'limi: barcha menejerlar allaqachon mavjud.")
        return

    print(f"  8-hafta bo'limi: {len(missing)} ta menejer qo'shilmoqda...")

    N = len(missing)
    old_last_8hf     = second_jami - 1
    last_existing_ri = old_last_8hf - tbl2_first

    ws.insert_rows(second_jami, amount=N)
    new_second_jami = second_jami + N
    new_last_8hf    = second_jami + N - 1

    def sum8(name):
        t = 0
        for d in dasturlar:
            m = next((x for x in d["managers"] if x["name"] == name), None)
            if m:
                t += sum(m["weeks"][:min(WK_LIMIT, len(m["weeks"]))])
        return t

    for ri, name in enumerate(missing):
        row = second_jami + ri
        ws.row_dimensions[row].height = 17
        bg  = BG if (last_existing_ri + 1 + ri) % 2 == 0 else ROW_ALT
        t8  = sum8(name)

        # A: Menejer (dark navy)
        c = ws.cell(row, 1, name)
        c.fill = fill(NAV_BG); c.font = font(bold=True, color=NAV_TXT, size=10)
        c.alignment = al("left"); c.border = border()

        # B: Jami (1-8 hafta)
        c = ws.cell(row, 2, t8 if t8 else None)
        c.fill = fill(HF_BG if t8 else bg)
        c.font = font(bold=True, color=HF_TXT if t8 else MUTED, size=11)
        c.alignment = al("center"); c.border = border(); c.number_format = "0"

        # C-I: har bir dastur uchun 1-8 hafta yig'indisi
        for di, d in enumerate(dasturlar):
            col = 3 + di
            m   = next((x for x in d["managers"] if x["name"] == name), None)
            v   = sum(m["weeks"][:min(WK_LIMIT, len(m["weeks"]))]) if m else 0
            dc_bg, dc_txt = D_COLORS.get(d["num"], ("DBEAFE", "1E40AF"))

            c = ws.cell(row, col)
            if v:
                c.value = v
                c.fill  = fill(dc_bg)
                c.font  = font(bold=True, color=dc_txt, size=10)
            else:
                c.value = None
                c.fill  = fill(bg)
                c.font  = font(color="CBD5E1", size=9)
            c.alignment = al("center"); c.border = border(); c.number_format = "0"

    # 8-hafta JAMI formulalarini qayta yozamiz
    fr2 = tbl2_first; lr2 = new_last_8hf
    for col in [2] + list(range(3, 3 + n_d)):
        cl = get_column_letter(col)
        c  = ws.cell(new_second_jami, col, f"=SUM({cl}{fr2}:{cl}{lr2})")
        c.number_format = "0"

    print(f"  8-hafta bo'limi: {N} ta menejer qo'shildi, JAMI yangilandi.")


# ── MAIN ─────────────────────────────────────────────────────────────────
print("CSV yuklanmoqda...")
dasturlar = fetch()
all_managers_new = sorted({m["name"] for d in dasturlar for m in d["managers"]})
for d in dasturlar:
    print(f"  {d['name']}: {len(d['weeks'])} hafta, {len(d['managers'])} menejer")
print(f"CSV da jami: {len(all_managers_new)} menejer")

print("\nMavjud Excel ochilmoqda...")
wb = load_workbook(EXCEL_PATH)
ws_hf = wb["HAFTALIK"]

# Asosiy jadval: yangi menejerlarni topamiz
existing, _ = get_sheet_managers(ws_hf, first_data_row=4)
print(f"Excel asosiy jadvalda mavjud: {len(existing)} menejer")

new_managers = sorted([m for m in all_managers_new if m not in existing])
changed = False

if new_managers:
    print(f"\nYangi menejerlar ({len(new_managers)} ta):")
    for m in new_managers:
        print(f"  + {m}")
    print("\nAsosiy jadvallar yangilanmoqda...")
    add_to_haftalik(ws_hf, dasturlar, new_managers, all_managers_new)
    if "UMUMIY JADVAL" in wb.sheetnames:
        add_to_umumiy(wb["UMUMIY JADVAL"], dasturlar, new_managers)
    pivot_name = "HAFTA × DASTUR × MENEJER"
    if pivot_name in wb.sheetnames:
        add_to_pivot(wb[pivot_name], dasturlar, new_managers)
    for d in dasturlar:
        if d["name"] in wb.sheetnames:
            new_in_d = [n for n in new_managers if any(m["name"]==n for m in d["managers"])]
            add_to_dastur(wb[d["name"]], d, new_in_d)
    changed = True
else:
    print("Asosiy jadvalda yangi menejer yo'q.")

# 8-hafta bo'limini har doim sinxronlashtir (asosiy jadvaldan mustaqil)
print("\n8-hafta bo'limi tekshirilmoqda...")
sync_8hafta_section(ws_hf, dasturlar, all_managers_new)
changed = True

wb.save(EXCEL_PATH)
print(f"\n✓ Saqlandi: {EXCEL_PATH}")
